import os
import time
import json
import random
import urllib.parse
import feedparser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google import genai

# ==========================================
# CONFIGURATION & SETTINGS
# ==========================================
OUTPUT_TXT_FILE = "rxbenefits_intel_hub_report.txt"
OUTPUT_EXCEL_FILE = "rxbenefits_intel_hub_report.xlsx"
OUTPUT_JSON_FILE = "daily_report.json"

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY environment variable is missing.")

AI_CLIENT = genai.Client(api_key=GEMINI_API_KEY)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

KEYWORDS = [
    "CVS Caremark", "OptumRx", "MedImpact", "Employers Health", "United Healthcare (UHC)",
    "Aetna", "Cigna Healthcare", "BCBS Affiliates", "BCBS TX", "BCBS AL", "BCBS IL", "BCBS NC",
    "Highmark BCBS", "BCBS MI", "Capital Blue Cross (Capital BCBS)", "SmithRx", "MedOne Rx",
    "Rightway Rx", "Aon Rx Coalition", "CoreTrust", "TrueScripts", "Liviniti", "Mercer Rx Coalition",
    "ProAct Rx", "Keenan Pharmacy Purchasing Coalition", "AffirmedRx", "Maxor", "VytlOne",
    "EmpiRx Health", "Innovative Rx Strategies", "National CooperativeRx", "NFP", "SlateRx",
    "TrueRx", "Welldyne", "Capital Rx", "LucyRx", "Elevance Health", "CarelonRx",
    "Prime Therapeutics", "Navitus", "WTW Collaborative", "Health Action Council",
    "RxCare Alliance", "Any BCBS", "Crumdale Specialty", "Ringmaster Technologies",
    "Risk International", "Rx-DNA", "S&S Consulting", "Truveris", "CVS Health", "Optum",
    "Judi Health", "Evernorth", "Express Scripts", "Meritain Health", "Carelon",
    "Magellan Health", "Anthem", "UnitedHealth Group", "UnitedHealthcare",
    "Managed Healthcare Executive", "Becker’s Payer Issues", "CMS", "BioPharma Dive",
    "PhRMA", "FTC", "HHS", "Reuters", "KFF", "National Community Pharmacists Association (NCPA)",
    "Forvis Mazars", "Health News Illinois", "InsuranceNewsNet", "JD Supra", "Drug Store News",
    "Fierce Healthcare", "Business Wire", "Morningstar", "American Hospital Association (AHA)",
    "Forbes Healthcare", "Medical Economics", "Drug Channels", "Brooklyn Research",
    "BenefitsPRO", "Pharmacy Benefit Manager", "PBM", "Pharmacy Benefits", "Specialty Pharmacy",
    "Drug Pricing", "Pharmacy Costs", "Prescription Drugs", "Pharmacy Rebates",
    "Rebate Aggregator", "Employer Benefits", "Employee Benefits", "Self-Insured Employers",
    "Health Plans", "Commercial Insurance", "Biosimilars", "GLP-1", "Weight Loss Drugs",
    "Oncology", "Specialty Drugs", "Formulary", "Prior Authorization", "Step Therapy",
    "Drug Transparency", "PBM Reform", "Pharmacy Transparency", "Drug Discount Cards",
    "Copay Assistance", "340B Program", "Employer Coalitions", "Pharmacy Purchasing Coalition",
    "AI in Healthcare", "Healthcare Analytics", "Healthcare Technology", "Employer Healthcare Costs",
    "Net Pharmacy Spend", "Gross Pharmacy Spend", "Pharmacy Inflation", "Healthcare Litigation",
    "PBM Litigation", "Federal Drug Pricing", "Pharmacy Benefit Innovation"
]

SYSTEM_PROMPT = """
You are an editor creating news updates for the RxBenefits Intel Hub News Radar Report. For each news article provided, follow these instructions exactly:

1. Filter the News: Only process articles that meet all of the following criteria:
   - Relevant to RxBenefits (pharmacy benefits management, self-funded employers, employee benefits, healthcare cost management, drug pricing, PBMs, specialty drugs).
   - Published by a US-based source.
   - Exclude any articles related to: Medicare, Medicaid, CMS.
   - Exclude duplicate articles or articles covering the same event/story.
2. Only give news which is highly relevant to RxBenefits.
3. Generate a New Title: Create a new, detailed, and impactful title that:
   - Is based on both the original title and article content.
   - Clearly communicates the key business insight.
   - Is not copied from the original title.
   - Is written in Title Case (capitalize the first letter of every major word).
   - Can span up to 3 lines as one continuous title.
4. Write the Summary: Write one concise paragraph (maximum 5 lines) that:
   - Completely paraphrases the article.
   - Does not copy sentences from the source.
   - Does not mention the original news headline.
   - Does not begin with phrases such as: "This article...", "The article...", "According to the article...", "The report..."
   - Directly explains the news and includes: What happened, Why it matters, Key business implications, and Important takeaways.
   - Uses only the information provided in the article. Do not add assumptions, opinions, or external information.
5. Output Format:
   If the article DOES NOT pass the filtering criteria, respond ONLY with: SKIP
   
   If the article PASSES, return the output strictly formatted with three labeled tags as shown below:
   TITLE: <Detailed Title in Title Case>
   SUMMARY: <One-paragraph summary (maximum 5 lines)>
   SOURCE_LINE: Source: <Source Name>; <Publication Date>
"""

def fetch_all_news(keywords):
    all_articles = []
    total = len(keywords)
    print(f"Starting news retrieval for {total} keywords...\n")

    for index, kw in enumerate(keywords, 1):
        print(f"[{index}/{total}] Fetching news for: '{kw}'...")
        query = f'"{kw}" when:24h'
        encoded_query = urllib.parse.quote(query)
        rss_url = f"https://news.google.com/rss/search?q={encoded_query}&hl=en-US&gl=US&ceid=US:en"
        
        retries = 0
        max_retries = 3
        success = False
        
        while retries < max_retries and not success:
            try:
                feed = feedparser.parse(rss_url, request_headers=HEADERS)
                if hasattr(feed, 'status') and feed.status == 429:
                    retries += 1
                    wait_time = 30 * (2 ** (retries - 1))
                    print(f"  -> Rate limit hit (429). Retrying in {wait_time}s...")
                    time.sleep(wait_time)
                    continue

                for item in feed.entries:
                    source_info = item.get('source', {})
                    source_name = source_info.get('title', 'N/A') if isinstance(source_info, dict) else getattr(source_info, 'title', 'N/A')

                    all_articles.append({
                        "keyword": kw,
                        "title": item.get('title', 'N/A'),
                        "link": item.get('link', 'N/A'),
                        "published": item.get('published', item.get('pubDate', 'N/A')),
                        "source_name": source_name,
                        "description": item.get('summary', item.get('description', 'N/A'))
                    })
                success = True
            except Exception as e:
                print(f"  -> Error fetching '{kw}': {e}")
                break

        time.sleep(random.uniform(1.5, 3.0))

    return all_articles

def process_article_with_ai(article):
    user_prompt = f"""
    Please evaluate and reformat this news article according to the rules:
    - Original Title: {article['title']}
    - Source Name: {article['source_name']}
    - Publication Date: {article['published']}
    - Link: {article['link']}
    - Content/Snippet: {article['description']}
    """

    try:
        response = AI_CLIENT.models.generate_content(
            model="gemini-2.5-flash",
            contents=f"{SYSTEM_PROMPT}\n\n{user_prompt}"
        )
        text = response.text.strip()
        if text.startswith("SKIP"):
            return None

        lines = text.split("\n")
        parsed = {"title": "", "summary": "", "source_line": "", "link": article['link'], "source_name": article['source_name'], "date": article['published']}
        
        for line in lines:
            if line.startswith("TITLE:"):
                parsed["title"] = line.replace("TITLE:", "").strip()
            elif line.startswith("SUMMARY:"):
                parsed["summary"] = line.replace("SUMMARY:", "").strip()
            elif line.startswith("SOURCE_LINE:"):
                parsed["source_line"] = line.replace("SOURCE_LINE:", "").strip()

        if not parsed["title"]:
            parsed["title"] = lines[0] if len(lines) > 0 else article['title']
        if not parsed["summary"]:
            parsed["summary"] = lines[1] if len(lines) > 1 else article['description']
        if not parsed["source_line"]:
            parsed["source_line"] = f"Source: {article['source_name']}; {article['published']}"

        return parsed
    except Exception as e:
        print(f"  -> AI Processing Error: {e}")
        return None

def save_json_report(processed_articles):
    with open(OUTPUT_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(processed_articles, f, indent=2)
    print(f"-> JSON Web data saved to: {os.path.abspath(OUTPUT_JSON_FILE)}")

def save_excel_format(processed_articles):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "News Radar Report"

    headers = ["New Title", "Summary", "Source Name", "Publication Date", "Hyperlink URL"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                         right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'),
                         bottom=Side(style='thin', color='D9D9D9'))

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, art in enumerate(processed_articles, start=2):
        ws.append([art["title"], art["summary"], art["source_name"], art["date"], art["link"]])
        ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            
        link_cell = ws.cell(row=row_idx, column=5)
        link_cell.font = Font(color="0000FF", underline="single")

    widths = [40, 60, 20, 25, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_EXCEL_FILE)

if __name__ == "__main__":
    raw_articles = fetch_all_news(KEYWORDS)
    processed_articles = []
    
    for idx, art in enumerate(raw_articles, 1):
        ai_output = process_article_with_ai(art)
        if ai_output:
            processed_articles.append(ai_output)
        time.sleep(0.3)

    save_json_report(processed_articles)
    if processed_articles:
        save_excel_format(processed_articles)
    print("Report pipeline finished successfully.")
